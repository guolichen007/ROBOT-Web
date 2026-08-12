from __future__ import annotations

import html
from datetime import UTC, datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy import select

from app.core.config import get_settings
from app.core.dependencies import AuthContext, DbSession, require_permission
from app.core.errors import PlatformError
from app.core.serialization import serialize_model
from app.db.models import InspectionObservation, PatrolReport, Robot, Task, TaskEvent

router = APIRouter(prefix="/api/v1/patrol-reports", tags=["patrol-reports"])


def report_root() -> Path:
    root = (get_settings().asset_root / "private" / "patrol-reports").resolve()
    root.mkdir(parents=True, exist_ok=True)
    return root


def safe_file(name: str) -> Path:
    root = report_root()
    target = (root / name).resolve()
    if root not in target.parents:
        raise PlatformError("INVALID_REPORT_PATH", "非法报告路径")
    return target


@router.get("")
def list_reports(
    db: DbSession, _: AuthContext = Depends(require_permission("robot.read"))
) -> list[dict]:
    return [
        serialize_model(row)
        for row in db.scalars(select(PatrolReport).order_by(PatrolReport.created_at.desc())).all()
    ]


@router.post("/tasks/{task_id}", status_code=201)
def generate_report(
    task_id: str,
    db: DbSession,
    _: AuthContext = Depends(require_permission("robot.read")),
) -> dict:
    from openpyxl import Workbook
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    task = db.get(Task, task_id)
    if not task or task.type != "PATROL":
        raise PlatformError("RESOURCE_NOT_FOUND", "巡检任务不存在", status_code=404)
    existing = db.scalar(select(PatrolReport).where(PatrolReport.task_id == task.id))
    if existing and existing.status == "READY":
        return serialize_model(existing)
    row = existing or PatrolReport(
        report_code=f"PR{datetime.now(UTC):%Y%m%d%H%M%S}-{task.id[:8]}", task_id=task.id
    )
    db.add(row)
    db.flush()
    robot = db.get(Robot, task.robot_id)
    observations = db.scalars(
        select(InspectionObservation)
        .where(InspectionObservation.task_id == task.id)
        .order_by(InspectionObservation.server_received_at)
    ).all()
    events = db.scalars(
        select(TaskEvent).where(TaskEvent.task_id == task.id).order_by(TaskEvent.created_at)
    ).all()
    summary = {
        "report_code": row.report_code,
        "task_code": task.task_code,
        "vehicle_id": robot.vehicle_id if robot else None,
        "map_version": task.map_version_snapshot,
        "status": task.status,
        "started_at": task.started_at.isoformat() if task.started_at else None,
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        "observation_total": len(observations),
        "abnormal_total": sum(item.result != "NORMAL" for item in observations),
        "generated_at": datetime.now(UTC).isoformat(),
    }
    prefix = row.report_code.lower()
    html_name, pdf_name, xlsx_name = f"{prefix}.html", f"{prefix}.pdf", f"{prefix}.xlsx"
    lines = (
        "".join(
            f"<tr><td>{html.escape(item.observation_type)}</td><td>{html.escape(item.result)}</td>"
            f"<td>{html.escape(item.data_state)}</td><td>{html.escape(str(item.value_json))}</td></tr>"
            for item in observations
        )
        or '<tr><td colspan="4">本次任务暂无传感器观测记录</td></tr>'
    )
    document = f"""<!doctype html><html lang='zh-CN'><meta charset='utf-8'>
    <style>body{{font-family:'Noto Sans CJK SC',sans-serif;color:#1f2933;margin:32px}}
    h1{{font-size:24px}} .meta{{display:grid;grid-template-columns:repeat(2,1fr);gap:8px}}
    table{{border-collapse:collapse;width:100%;margin-top:20px}}
    th,td{{border:1px solid #cbd2d9;padding:8px}}
    th{{background:#edf1f4;text-align:left}}</style>
    <h1>智能灭火机器人巡检报告</h1><div class='meta'>
    <div>报告编号：{row.report_code}</div><div>任务编号：{task.task_code}</div>
    <div>车辆：{summary["vehicle_id"]}</div><div>地图版本：{task.map_version_snapshot}</div>
    <div>任务状态：{task.status}</div><div>异常数：{summary["abnormal_total"]}</div></div>
    <table><thead><tr><th>观测项</th><th>结果</th><th>数据状态</th><th>值</th></tr></thead>
    <tbody>{lines}</tbody></table><p>生成时间（UTC）：{summary["generated_at"]}</p></html>"""
    safe_file(html_name).write_text(document, encoding="utf-8")
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Normal"):
        styles[style_name].fontName = "STSong-Light"
    pdf_rows = [
        ["报告编号", row.report_code],
        ["任务编号", task.task_code],
        ["车辆", str(summary["vehicle_id"] or "--")],
        ["地图版本", task.map_version_snapshot],
        ["任务状态", task.status],
        ["异常数", str(summary["abnormal_total"])],
    ]
    details = [["观测项", "结果", "数据状态", "值"]] + [
        [item.observation_type, item.result, item.data_state, str(item.value_json)]
        for item in observations
    ]
    if len(details) == 1:
        details.append(["本次任务暂无传感器观测记录", "--", "--", "--"])
    meta_table = Table(pdf_rows, colWidths=[90, 350])
    detail_table = Table(details, repeatRows=1, colWidths=[100, 70, 90, 180])
    common_style = TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd2d9")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#edf1f4")),
            ("PADDING", (0, 0), (-1, -1), 7),
        ]
    )
    meta_table.setStyle(common_style)
    detail_table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd2d9")),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#edf1f4")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    pdf = SimpleDocTemplate(str(safe_file(pdf_name)), pagesize=A4)
    pdf.build(
        [
            Paragraph("智能灭火机器人巡检报告", styles["Title"]),
            Spacer(1, 14),
            meta_table,
            Spacer(1, 18),
            detail_table,
            Spacer(1, 12),
            Paragraph(f"生成时间（UTC）：{summary['generated_at']}", styles["Normal"]),
        ]
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "巡检汇总"
    sheet.append(["报告编号", row.report_code])
    sheet.append(["任务编号", task.task_code])
    sheet.append(["车辆", summary["vehicle_id"]])
    sheet.append(["地图版本", task.map_version_snapshot])
    sheet.append(["任务状态", task.status])
    detail = workbook.create_sheet("观测明细")
    detail.append(["观测项", "结果", "数据状态", "源时间", "接收时间", "值"])
    for item in observations:
        detail.append(
            [
                item.observation_type,
                item.result,
                item.data_state,
                item.source_timestamp.isoformat() if item.source_timestamp else None,
                item.server_received_at.isoformat(),
                str(item.value_json),
            ]
        )
    timeline = workbook.create_sheet("任务时间线")
    timeline.append(["状态", "阶段", "进度", "时间"])
    for event in events:
        timeline.append([event.status, event.phase, event.progress, event.created_at.isoformat()])
    workbook.save(safe_file(xlsx_name))
    row.status = "READY"
    row.summary_json = summary
    row.html_object_name = html_name
    row.pdf_object_name = pdf_name
    row.xlsx_object_name = xlsx_name
    row.generated_at = datetime.now(UTC)
    db.commit()
    return serialize_model(row)


@router.get("/{report_id}/download/{format_name}")
def download_report(
    report_id: str,
    format_name: str,
    db: DbSession,
    _: AuthContext = Depends(require_permission("robot.read")),
) -> FileResponse:
    row = db.get(PatrolReport, report_id)
    if not row or row.status != "READY":
        raise PlatformError("RESOURCE_NOT_FOUND", "报告尚未生成", status_code=404)
    mapping = {
        "html": (row.html_object_name, "text/html; charset=utf-8"),
        "pdf": (row.pdf_object_name, "application/pdf"),
        "xlsx": (
            row.xlsx_object_name,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }
    if format_name not in mapping:
        raise PlatformError("FORMAT_UNSUPPORTED", "只支持 html/pdf/xlsx")
    name, media_type = mapping[format_name]
    if not name or not safe_file(name).exists():
        raise PlatformError("RESOURCE_NOT_FOUND", "报告文件不存在", status_code=404)
    return FileResponse(safe_file(name), media_type=media_type, filename=name)
