from __future__ import annotations

import json
from datetime import UTC, datetime
from uuid import uuid4

import pytest
from app.core.config import get_settings
from app.core.events import append_event, get_redis
from app.core.security import hash_password, verify_password
from app.db.models import (
    Command,
    FireEvent,
    Map,
    MapVersion,
    OutboxEvent,
    ParkingSlot,
    Robot,
    RobotCapability,
    RobotIntegrationProfile,
    RobotMotionProfile,
    Task,
    User,
)
from app.db.session import SessionLocal
from app.main import app
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select, update
from starlette.websockets import WebSocketDisconnect


@pytest.fixture(autouse=True)
def stable_baseline() -> None:
    get_redis().flushdb()
    with SessionLocal.begin() as db:
        admin = db.scalar(select(User).where(User.username == "admin"))
        assert admin
        test_password = get_settings().effective_admin_password
        if not verify_password(test_password, admin.password_hash):
            admin.password_hash = hash_password(test_password)
        admin.must_change_password = False
        admin.failed_attempts = 0
        admin.locked_until = None
        robot = db.scalar(select(Robot).where(Robot.vehicle_id == "R001"))
        assert robot
        active_map = db.scalar(select(Map).where(Map.active_version_id.is_not(None)))
        assert active_map
        version = db.get(MapVersion, active_map.active_version_id)
        assert version
        robot.online_state = "ONLINE"
        robot.boot_id = "00000000-0000-4000-8000-000000000002"
        robot.last_seen_at = datetime.now(UTC)
        robot.estop_active = False
        robot.current_map_id = active_map.id
        robot.current_map_version = version.version
        integration = db.get(RobotIntegrationProfile, robot.id)
        if integration:
            integration.source_kind = "MOCK"
            integration.control_contract_verified = True
            integration.ack_contract_verified = True
            integration.map_contract_verified = True
            integration.bidirectional_bridge_verified = True
            integration.command_path_verified = True
            integration.cmd_vel_arbitration_verified = True
            integration.read_only_reason = None
        capability = db.get(RobotCapability, robot.id)
        if capability:
            capability.supported_commands_json = [
                "manual_control",
                "stop_motion",
                "emergency_stop",
                "reset_estop",
                "return_dock",
                "patrol",
                "extinguish",
                "cancel_task",
            ]
        motion = db.get(RobotMotionProfile, robot.id)
        if motion:
            motion.manual_watchdog_verified = True
            motion.max_manual_forward_mps = 0.22
            motion.max_manual_reverse_mps = 0.16
            motion.max_manual_angular_radps = 0.65
            motion.reverse_allowed = True
        db.execute(
            update(Task)
            .where(Task.status.in_({"CREATED", "QUEUED", "ACCEPTED", "EXECUTING"}))
            .values(status="SUCCEEDED", phase="TEST_CLEANUP", completed_at=datetime.now(UTC))
        )


@pytest.fixture
def client() -> TestClient:
    with TestClient(app) as value:
        yield value


def login(client: TestClient) -> str:
    response = client.post(
        "/api/v1/auth/login",
        json={"username": "admin", "password": get_settings().effective_admin_password},
    )
    assert response.status_code == 200, response.text
    return str(response.json()["access_token"])


def auth(token: str, **headers: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}", **headers}


def target() -> tuple[Robot, ParkingSlot, MapVersion]:
    with SessionLocal() as db:
        robot = db.scalar(select(Robot).where(Robot.vehicle_id == "R001"))
        slot = db.scalar(select(ParkingSlot).where(ParkingSlot.code == "A-12"))
        assert robot and slot
        version = db.get(MapVersion, slot.map_version_id)
        assert version
        db.expunge(robot)
        db.expunge(slot)
        db.expunge(version)
        return robot, slot, version


def admin_id() -> str:
    with SessionLocal() as db:
        admin = db.scalar(select(User).where(User.username == "admin"))
        assert admin
        return admin.id


def test_refresh_rotation_reuse_revokes_family(client: TestClient) -> None:
    login(client)
    old_refresh = client.cookies.get("refresh_token")
    old_csrf = client.cookies.get("csrf_token")
    assert old_refresh and old_csrf

    rotated = client.post("/api/v1/auth/refresh", headers={"X-CSRF-Token": old_csrf})
    assert rotated.status_code == 200
    new_refresh = client.cookies.get("refresh_token")
    new_csrf = client.cookies.get("csrf_token")
    assert new_refresh and new_refresh != old_refresh and new_csrf

    with TestClient(app) as replay:
        replay.cookies.set("refresh_token", old_refresh)
        replay.cookies.set("csrf_token", old_csrf)
        rejected = replay.post("/api/v1/auth/refresh", headers={"X-CSRF-Token": old_csrf})
        assert rejected.status_code == 401

    revoked = client.post("/api/v1/auth/refresh", headers={"X-CSRF-Token": new_csrf})
    assert revoked.status_code == 401


def test_manual_lease_mutex_sequence_and_release(client: TestClient) -> None:
    token = login(client)
    first = client.post(
        "/api/v1/robots/R001/manual-lease",
        json={"control_session_id": str(uuid4())},
        headers=auth(token),
    )
    assert first.status_code == 201
    lease = first.json()

    second = client.post(
        "/api/v1/robots/R001/manual-lease",
        json={"control_session_id": str(uuid4())},
        headers=auth(token),
    )
    assert second.status_code == 409

    pulse = {
        "lease_id": lease["lease_id"],
        "control_session_id": lease["control_session_id"],
        "seq": 1,
        "linear": 0.2,
        "angular": 0,
    }
    accepted = client.post("/api/v1/robots/R001/commands/manual", json=pulse, headers=auth(token))
    assert accepted.status_code == 200 and accepted.json()["accepted"] is True
    duplicate = client.post("/api/v1/robots/R001/commands/manual", json=pulse, headers=auth(token))
    assert duplicate.status_code == 200
    assert duplicate.json() == {
        "accepted": False,
        "reason": "DUPLICATE_OR_OUT_OF_ORDER",
        "last_seq": 1,
    }

    released = client.delete("/api/v1/robots/R001/manual-lease", headers=auth(token))
    assert released.status_code == 204


def test_manual_command_is_clamped_by_server_motion_profile(client: TestClient) -> None:
    token = login(client)
    lease = client.post(
        "/api/v1/robots/R001/manual-lease",
        json={"control_session_id": str(uuid4())},
        headers=auth(token),
    )
    assert lease.status_code == 201
    held = lease.json()
    command = client.post(
        "/api/v1/robots/R001/commands/manual",
        json={
            "lease_id": held["lease_id"],
            "control_session_id": held["control_session_id"],
            "seq": 1,
            "linear": 0.4,
            "angular": 1.0,
        },
        headers=auth(token),
    )
    assert command.status_code == 200
    assert command.json()["clamped"] is True
    assert command.json()["applied"] == {"linear": 0.22, "angular": 0.65}
    client.delete("/api/v1/robots/R001/manual-lease", headers=auth(token))


def test_stale_pose_never_draws_detection_coverage(client: TestClient) -> None:
    token = login(client)
    get_redis().set(
        "robot:R001:latest",
        json.dumps(
            {
                "x": 10,
                "y": 10,
                "theta": 0,
                "localization_status": "GOOD",
                "server_received_at": "2020-01-01T00:00:00+00:00",
            }
        ),
    )
    response = client.get("/api/v1/robots/R001/detection-coverage", headers=auth(token))
    assert response.status_code == 200
    assert response.json()["state"] == "STALE"
    assert response.json()["polygon"] == []
    assert response.json()["covered_parking_slot_ids"] == []


def test_manual_lease_blocks_autonomous_task(client: TestClient) -> None:
    token = login(client)
    lease = client.post(
        "/api/v1/robots/R001/manual-lease",
        json={"control_session_id": str(uuid4())},
        headers=auth(token),
    )
    assert lease.status_code == 201
    _, slot, _ = target()
    response = client.post(
        "/api/v1/tasks/patrol",
        json={"robot_id": "R001", "target_parking_slot_id": slot.id},
        headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
    )
    assert response.status_code == 409
    client.delete("/api/v1/robots/R001/manual-lease", headers=auth(token))


def test_logout_releases_lease_and_queues_stop(client: TestClient) -> None:
    token = login(client)
    robot, _, _ = target()
    lease = client.post(
        "/api/v1/robots/R001/manual-lease",
        json={"control_session_id": str(uuid4())},
        headers=auth(token),
    )
    assert lease.status_code == 201
    csrf = client.cookies.get("csrf_token")
    assert csrf
    logged_out = client.post("/api/v1/auth/logout", headers={"X-CSRF-Token": csrf})
    assert logged_out.status_code == 204
    assert not get_redis().exists(f"manual:lease:{robot.id}")
    with SessionLocal() as db:
        stop = db.scalar(
            select(Command)
            .where(Command.robot_id == robot.id, Command.cmd == "stop_motion")
            .order_by(Command.issued_at.desc())
        )
        assert stop and stop.payload_json["params"]["reason"] == "USER_LOGOUT"


def test_task_idempotency_outbox_and_map_mismatch(client: TestClient) -> None:
    token = login(client)
    robot, slot, _ = target()
    key = str(uuid4())
    payload = {"robot_id": "R001", "target_parking_slot_id": slot.id}
    first = client.post(
        "/api/v1/tasks/patrol",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": key}),
    )
    assert first.status_code == 201
    replay = client.post(
        "/api/v1/tasks/patrol",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": key}),
    )
    assert replay.status_code == 201
    assert replay.json()["id"] == first.json()["id"]
    with SessionLocal() as db:
        command = db.scalar(select(Command).where(Command.command_id == first.json()["command_id"]))
        assert command and command.lifecycle_status == "QUEUED"
        assert db.scalar(select(OutboxEvent).where(OutboxEvent.aggregate_id == command.command_id))

    with SessionLocal.begin() as db:
        current = db.get(Robot, robot.id)
        assert current
        current.current_map_version = "MISMATCH"
        db.execute(update(Task).where(Task.id == first.json()["id"]).values(status="SUCCEEDED"))
    mismatch = client.post(
        "/api/v1/tasks/patrol",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
    )
    assert mismatch.status_code == 409
    assert mismatch.json()["error"]["code"] == "MAP_VERSION_MISMATCH"
    assert mismatch.json()["error"]["details"]["robot_map_version"] == "MISMATCH"


def test_report_download_requires_bearer_and_returns_real_pdf_xlsx(
    client: TestClient, tmp_path
) -> None:
    token = login(client)
    robot, slot, version = target()
    task = Task(
        task_code=f"REPORT-{uuid4()}",
        robot_id=robot.id,
        type="PATROL",
        status="SUCCEEDED",
        phase="COMPLETED",
        progress=100,
        target_parking_slot_id=slot.id,
        target_pose_snapshot_json=slot.center_pose_json,
        map_id_snapshot=version.map_id,
        map_version_snapshot=version.version,
        semantic_revision_snapshot=version.semantic_revision,
        trajectory_snapshot_json=[],
        parameters_json={"source": "TEST"},
        created_by=admin_id(),
        started_at=datetime.now(UTC),
        completed_at=datetime.now(UTC),
    )
    with SessionLocal.begin() as db:
        db.add(task)
    generated = client.post(f"/api/v1/patrol-reports/tasks/{task.id}", headers=auth(token))
    assert generated.status_code == 201, generated.text
    report = generated.json()
    assert report["pdf_asset_id"] and report["xlsx_asset_id"]
    for format_name in ("pdf", "xlsx"):
        unauthenticated = client.get(
            f"/api/v1/patrol-reports/{report['id']}/download/{format_name}"
        )
        assert unauthenticated.status_code == 401
    pdf = client.get(f"/api/v1/patrol-reports/{report['id']}/download/pdf", headers=auth(token))
    assert pdf.status_code == 200 and pdf.content.startswith(b"%PDF-")
    xlsx = client.get(f"/api/v1/patrol-reports/{report['id']}/download/xlsx", headers=auth(token))
    assert xlsx.status_code == 200 and xlsx.content.startswith(b"PK")
    xlsx_path = tmp_path / "report.xlsx"
    xlsx_path.write_bytes(xlsx.content)
    workbook = load_workbook(xlsx_path, read_only=True)
    assert {"巡检汇总", "观测明细", "任务时间线"}.issubset(workbook.sheetnames)


def test_manual_alarm_idempotency_lifecycle_and_audit(client: TestClient) -> None:
    token = login(client)
    _, slot, version = target()
    key = str(uuid4())
    payload = {
        "parking_slot_id": slot.id,
        "fire_type": "smoke",
        "note": "integration test",
        "map_version": version.version,
    }
    created = client.post(
        "/api/v1/alarms/manual",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": key}),
    )
    assert created.status_code == 201
    replay = client.post(
        "/api/v1/alarms/manual",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": key}),
    )
    assert replay.status_code == 201 and replay.json()["id"] == created.json()["id"]
    alarm_id = created.json()["id"]
    assert created.json()["state"] == "NEW"
    timeline = client.get(f"/api/v1/alarms/{alarm_id}/timeline", headers=auth(token))
    assert timeline.status_code == 200
    assert [(item["source_type"], item["state"]) for item in timeline.json()] == [("ALARM", "NEW")]
    for action, state in (
        ("acknowledge", "ACKNOWLEDGED"),
        ("confirm", "CONFIRMED"),
        ("resolve", "RESOLVED"),
    ):
        transitioned = client.post(f"/api/v1/alarms/{alarm_id}/{action}", headers=auth(token))
        assert transitioned.status_code == 200 and transitioned.json()["state"] == state
    with SessionLocal() as db:
        assert db.get(FireEvent, alarm_id).state == "RESOLVED"  # type: ignore[union-attr]


def test_extinguish_task_allowed_before_confirm_rejected_after_resolve(client: TestClient) -> None:
    token = login(client)
    _, slot, version = target()

    def create_alarm(note: str) -> str:
        created = client.post(
            "/api/v1/alarms/manual",
            json={
                "parking_slot_id": slot.id,
                "fire_type": "smoke",
                "note": note,
                "map_version": version.version,
            },
            headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
        )
        assert created.status_code == 201, created.text
        return str(created.json()["id"])

    payload = {
        "robot_id": "R001",
        "parameters": {"source": "OPERATIONS_HMI", "extinguish_mode": "DEPLOY_BLANKET"},
    }

    # NEW alarm: direct extinguish dispatch is allowed without prior confirmation
    new_id = create_alarm("direct extinguish new")
    created_task = client.post(
        f"/api/v1/alarms/{new_id}/create-task",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
    )
    assert created_task.status_code == 201, created_task.text

    # RESOLVED alarm: dispatch is rejected even with authorization
    resolved_id = create_alarm("direct extinguish resolved")
    transitioned = client.post(f"/api/v1/alarms/{resolved_id}/resolve", headers=auth(token))
    assert transitioned.status_code == 200, transitioned.text
    rejected = client.post(
        f"/api/v1/alarms/{resolved_id}/create-task",
        json=payload,
        headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
    )
    assert rejected.status_code == 409


def test_snapshot_watermark_replay_and_one_time_ws_ticket(client: TestClient) -> None:
    token = login(client)
    snapshot = client.get("/api/v1/monitor/snapshot", headers=auth(token))
    assert snapshot.status_code == 200
    watermark = snapshot.json()["snapshot_watermark"]
    emitted = append_event("vehicle.location", {"vehicle_id": "R001", "x": 9, "y": 8})
    ticket = client.post("/api/v1/auth/ws-ticket", headers=auth(token)).json()["ticket"]
    with client.websocket_connect(
        f"/ws/v1/monitor?ticket={ticket}&after={watermark}",
    ) as websocket:
        replayed = []
        for _ in range(20):
            event = websocket.receive_json()
            replayed.append(event)
            if event["stream_id"] == emitted:
                break
        assert any(item["stream_id"] == emitted for item in replayed)
        assert next(item for item in replayed if item["stream_id"] == emitted)["data"]["x"] == 9
    with pytest.raises(WebSocketDisconnect):
        with client.websocket_connect(
            f"/ws/v1/monitor?ticket={ticket}&after={watermark}",
        ):
            pass


def test_health_ready_and_offline_estop_semantics(client: TestClient) -> None:
    token = login(client)
    ready = client.get("/health/ready")
    assert ready.status_code == 503 and ready.json()["ok"] is False
    with SessionLocal.begin() as db:
        robot = db.scalar(select(Robot).where(Robot.vehicle_id == "R001"))
        assert robot
        robot.online_state = "OFFLINE"
    command = client.post(
        "/api/v1/robots/R001/commands/emergency-stop",
        json={},
        headers=auth(token, **{"Idempotency-Key": str(uuid4())}),
    )
    assert command.status_code == 202
    assert command.json()["lifecycle_status"] == "PUBLISHED_UNCONFIRMED"
    assert command.json()["ack_reason"] == "OFFLINE_NOT_DELIVERED"
